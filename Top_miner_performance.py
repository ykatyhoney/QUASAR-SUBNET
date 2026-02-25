#!/usr/bin/env python3
"""
Check miner submissions and rankings.
Shows top 4 miner hotkeys (including round winner #1) for the latest COMPLETED round only.
Supports JSON and XLS export.
"""
import argparse
import json
import os
import sys
import requests

API_URL = os.environ.get("QUASAR_API_URL", "http://localhost:8000")
NETWORK = os.environ.get("QUASAR_NETWORK", "finney")


def main():
    parser = argparse.ArgumentParser(
        description="Top 4 miners for latest completed round. Optional JSON/XLS export."
    )
    parser.add_argument("--json", metavar="FILE", help="Write JSON output to FILE")
    parser.add_argument("--xls", metavar="FILE", help="Write XLS (.xlsx) output to FILE (requires openpyxl)")
    parser.add_argument("--quiet", action="store_true", help="Only output export files, no console summary")
    args = parser.parse_args()

    # 1. Get current round
    if not args.quiet:
        print("=" * 60)
        print("CURRENT ROUND")
        print("=" * 60)
    try:
        r = requests.get(
            f"{API_URL}/get_current_round",
            params={"network": NETWORK},
            timeout=15,
        )
        round_data = r.json() if r.text else {}
        if r.status_code != 200 or "detail" in round_data:
            if not args.quiet:
                print(round_data.get("detail", r.text or "No response"))
        elif not args.quiet:
            print(json.dumps(round_data, indent=2))
    except Exception as e:
        if not args.quiet:
            print(f"Error: {e}")
        round_data = {}

    # 2. Get submission stats
    if not args.quiet:
        print("\n" + "=" * 60)
        print("SUBMISSION STATISTICS")
        print("=" * 60)
    try:
        stats = requests.get(f"{API_URL}/get_submission_stats", timeout=15).json()
    except Exception:
        stats = {}
    if not args.quiet:
        print(f"Total Submissions: {stats.get('total_submissions', 0)}")
        s = stats.get("stats", {})
        print(f"Average Performance: {s.get('avg_tokens_per_sec', 0):.2f} tokens/sec")
        print(f"Max Performance: {s.get('max_tokens_per_sec', 0):.2f} tokens/sec")

    # 3. Top 10 performers
    if not args.quiet:
        print("\n" + "=" * 60)
        print("TOP 10 PERFORMERS (ALL TIME)")
        print("=" * 60)
        for i, miner in enumerate(stats.get("top_performers", [])[:10], 1):
            print(
                f"{i}. {miner.get('miner_hotkey', '')[:12]}... | "
                f"{miner.get('tokens_per_sec', 0):.0f} tokens/sec | "
                f"Seq Len: {miner.get('target_sequence_length', 'N/A')}"
            )

    # 4. Recent submissions
    if not args.quiet:
        print("\n" + "=" * 60)
        recent = stats.get("recent_submissions", [])
        print(f"RECENT SUBMISSIONS (Last {len(recent)})")
        print("=" * 60)
        for sub in recent[:10]:
            status = "Validated" if sub.get("validated") else "Pending"
            print(
                f"{sub.get('miner_hotkey', '')[:12]}... | "
                f"{sub.get('tokens_per_sec', 0):.0f} tokens/sec | "
                f"{status} | Round: {round_data.get('round_number', 'N/A')}"
            )

    # 5. Top 4 miners for LATEST COMPLETED ROUND only (never active round)
    if not args.quiet:
        print("\n" + "=" * 60)
        print("TOP 4 MINERS (LATEST COMPLETED ROUND ONLY)")
        print("=" * 60)
    try:
        r = requests.get(
            f"{API_URL}/get_weights",
            params={"network": NETWORK, "completed_only": True},
            timeout=15,
        )
        if not r.text or r.status_code != 200:
            weights_data = None
            wlist = []
            if not args.quiet:
                print("No completed round or API error.")
        else:
            weights_data = r.json()
            wlist = weights_data.get("weights") or []
    except requests.exceptions.JSONDecodeError:
        weights_data = None
        wlist = []
        if not args.quiet:
            print("API returned non-JSON (check server).")
    except Exception as e:
        weights_data = None
        wlist = []
        if not args.quiet:
            print(f"Error: {e}")

    if wlist:
        round_label = (
            (weights_data or {}).get("round_number")
            or (weights_data or {}).get("round_id")
            or "?"
        )
        round_status = (weights_data or {}).get("round_status", "")
        if not args.quiet:
            print(f"Round: {round_label} (status: {round_status})\n")
            winner = wlist[0].get("hotkey", "") if wlist else None
            if winner:
                print(f"Round winner (top 1): {winner[:12]}...\n")
            for i, w in enumerate(wlist[:4], 1):
                uid = w.get("uid", 0)
                hk = w.get("hotkey", "")
                weight = w.get("weight", 0)
                tps = w.get("tokens_per_sec")
                gh = w.get("github_username") or "-"
                short = (hk[:12] + "...") if len(hk) > 12 else hk
                tps_str = f" | {tps:.0f} tok/s" if tps is not None else ""
                print(
                    f"  #{i} UID {uid}: {short} | GitHub: {gh} | "
                    f"Weight: {weight:.2%}{tps_str}"
                )
    elif not args.quiet:
        print("No completed rounds yet.")

    # Build payload for JSON/XLS export
    payload = {
        "network": NETWORK,
        "api_url": API_URL,
        "current_round_number": round_data.get("round_number"),
        "latest_completed_round": {
            "round_id": (weights_data or {}).get("round_id"),
            "round_number": (weights_data or {}).get("round_number"),
            "round_status": (weights_data or {}).get("round_status"),
            "winner_hotkey": wlist[0].get("hotkey") if wlist else None,
        }
        if weights_data
        else None,
        "top4_miners": [
            {
                "rank": i,
                "is_winner": i == 1,
                "uid": w.get("uid"),
                "hotkey": w.get("hotkey"),
                "weight": w.get("weight"),
                "tokens_per_sec": w.get("tokens_per_sec"),
                "github_username": w.get("github_username"),
            }
            for i, w in enumerate(wlist[:4], 1)
        ],
    }

    if args.json:
        with open(args.json, "w") as f:
            json.dump(payload, f, indent=2)
        print(f"Wrote JSON: {args.json}", file=sys.stderr)

    if args.xls:
        try:
            import openpyxl
            from openpyxl.styles import Font

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Latest completed round"
            headers = [
                "Round", "Status", "Winner hotkey",
                "Rank", "UID", "Hotkey", "Weight", "Tokens/sec", "GitHub",
            ]
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
                ws.cell(row=1, column=col).font = Font(bold=True)
            rc = payload.get("latest_completed_round") or {}
            ws.cell(row=2, column=1, value=rc.get("round_number"))
            ws.cell(row=2, column=2, value=rc.get("round_status"))
            ws.cell(row=2, column=3, value=rc.get("winner_hotkey"))
            row = 3
            for m in payload.get("top4_miners") or []:
                ws.cell(row=row, column=4, value=m.get("rank"))
                ws.cell(row=row, column=5, value=m.get("uid"))
                ws.cell(row=row, column=6, value=m.get("hotkey"))
                ws.cell(row=row, column=7, value=m.get("weight"))
                ws.cell(row=row, column=8, value=m.get("tokens_per_sec"))
                ws.cell(row=row, column=9, value=m.get("github_username"))
                row += 1
            wb.save(args.xls)
            print(f"Wrote XLS: {args.xls}", file=sys.stderr)
        except ImportError:
            print("Install openpyxl for XLS: pip install openpyxl", file=sys.stderr)


if __name__ == "__main__":
    main()
